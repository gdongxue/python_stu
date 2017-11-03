# -*- coding:utf-8 -*-
import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
file = "C:/Users/dongxue/Desktop/test.xlsx"
inw = load_workbook(file)
a = Workbook()
sheetname = inw.get_sheet_by_name("three")
# rows= sheetname.max_row()
# print rows
print sheetname


b = Workbook()
c = b.create_sheet("gaodongxue")
c.cell(row=1,column=1).value
c.cell(row=2,column=2).value
b.save("gaodongxue.xlsx")